
from django.shortcuts import redirect
from .models import Teacher
from django.shortcuts import render, redirect, get_object_or_404
from .forms import TeacherForm
from attendance.models import Attendance
import csv
from openpyxl import load_workbook
from django.contrib.auth.models import User, Group
from django.contrib.auth import get_user_model
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils import timezone
from admission.models import StudentProfile
from django.contrib.auth.decorators import user_passes_test

User = get_user_model() 
def is_teacher(user):
    return user.groups.filter(name="Teacher").exists() or user.is_superuser

@login_required
@user_passes_test(is_teacher)
def teacher_dashboard(request):
    teacher = Teacher.objects.get(user=request.user)
    school = request.user.userprofile.school   # ✅ ONLY SOURCE
    today = timezone.now().date()

    students = StudentProfile.objects.filter(
        admission__class_applied=teacher.class_name,
        admission__section=teacher.section,
        school=school
    )

    total_students = students.count()

    present = Attendance.objects.filter(
        student__in=students,
        date=today,
        status="P"
    ).count()

    absent = total_students - present
    percent = round((present / total_students) * 100, 2) if total_students else 0

    return render(
        request,
        "teacher/teacher_dashboard.html",
        {
            "teacher": teacher,
            "total_students": total_students,
            "present": present,
            "absent": absent,
            "percent": percent,
            "school": school,   # optional for UI
        }
    )

def teacher_list(request):
    query = request.GET.get("subject")

    teachers = Teacher.objects.filter(
        user__school=request.user.school
    ).select_related("user")

    if query:
        teachers = teachers.filter(subject__icontains=query)

    return render(
        request,
        "teacher/teacher_list.html",
        {"teacher": teachers}
    )
@login_required
def teacher_edit(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)

    if request.method == "POST":
        form = TeacherForm(request.POST, instance=teacher)
        if form.is_valid():
            form.save()
            return redirect("teachers:list")
    else:
        form = TeacherForm(instance=teacher)

    return render(request, "teacher/teacher_form.html", {"form": form})


@login_required
def teacher_delete(request, pk):
    teacher = get_object_or_404(Teacher, pk=pk)
    teacher.delete()
    return redirect("teachers:list")


@login_required
def teacher_attendance(request):
    teacher = Teacher.objects.get(user=request.user)

    records = Attendance.objects.filter(
        student__class_name=teacher.class_name,
        student__section=teacher.section,
        student__school=request.user.school
    ).select_related("student").order_by("-date")

    return render(
        request,
        "teacher/teacher_attendance.html",
        {"records": records}
    )

@login_required
def mark_attendance(request):
    # 🔐 Sirf Teacher allow
    if not request.user.groups.filter(name="Teacher").exists():
        return redirect("schools:dashboard")

    teacher = Teacher.objects.get(user=request.user)
    today = timezone.now().date()

    students = StudentProfile.objects.filter(
        class_name=teacher.class_name,
        section=teacher.section
    )

    if request.method == "POST":
        for student in students:
            status = request.POST.get(str(student.id))  # "P" or "A"
            Attendance.objects.update_or_create(
                student=student,
                date=today,
                defaults={
                    "teacher": teacher,
                    "is_present": status == "P"
                }
            )
        return redirect("teachers:attendance")

    return render(
        request,
        "teacher/mark_attendance.html",
        {
            "students": students,
            "today": today
        }
    )

def is_school_admin(user):
    return user.is_superuser or user.groups.filter(name="SchoolAdmin").exists()

@user_passes_test(is_school_admin)

def import_teachers(request):
    if request.method == "POST":
        file = request.FILES.get("file")

        if not file:
            messages.error(request, "No file uploaded")
            return redirect("teachers:import")

        teacher_group, _ = Group.objects.get_or_create(name="Teacher")
        created_count = 0

        # CSV
        if file.name.endswith(".csv"):
            reader = csv.DictReader(
                file.read().decode("utf-8", errors="ignore").splitlines()
            )

        # Excel
        elif file.name.endswith(".xlsx"):
            wb = load_workbook(file)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            reader = (
                dict(zip(headers, row))
                for row in sheet.iter_rows(min_row=2, values_only=True)
            )
        else:
            messages.error(request, "Only CSV or Excel allowed")
            return redirect("teachers:import")

        for row in reader:
            user, created = User.objects.get_or_create(
                username=row["username"],
                defaults={
                    "email": row.get("email", ""),
                    "first_name": row.get("first_name", ""),
                    "last_name": row.get("last_name", ""),
                    
                }
            )

            if created:
                user.school = request.user.school
                user.set_password("Teacher@123")
                user.save()
                user.groups.add(teacher_group)

            # 🔥 NO school FIELD HERE
            teacher, t_created = Teacher.objects.get_or_create(
                user=user,
                defaults={
                    "class_name": row["class_name"],
                    "section": row["section"],
                }
            )

            if t_created:
                created_count += 1

        messages.success(
            request,
            f"{created_count} teachers imported successfully. Default password: Teacher@123"
        )

        return redirect("teachers:list")

    return render(request, "teacher/import_teachers.html")


def teacher_attendance_admin(request, teacher_id):
    teacher = Teacher.objects.get(id=teacher_id)

    records = Attendance.objects.filter(
        student__class_name=teacher.class_name,
        student__section=teacher.section,
        student__school=teacher.school
    )

    return render(
        request,
        "school/teacher_attendance_admin.html",
        {
            "teacher": teacher,
            "records": records
        }
    )
@login_required
def assign_subject(request):
    if request.method == "POST":
        subject = request.POST.get("subject")
        teacher_ids = request.POST.getlist("teachers")

        if not subject or not teacher_ids:
            messages.error(request, "Please select teachers and enter subject")
            return redirect("teachers:list")

        Teacher.objects.filter(
            id__in=teacher_ids,
            user__school=request.user.school
        ).update(subject=subject)

        messages.success(
            request,
            f"Subject '{subject}' assigned to {len(teacher_ids)} teachers"
        )

    return redirect("teachers:list")